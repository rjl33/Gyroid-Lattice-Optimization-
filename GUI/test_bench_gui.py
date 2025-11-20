"""
Test Bench GUI - Compression and Torsion Testing Machine
PySide6 application for Raspberry Pi
"""

import sys
import csv
import datetime
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QLineEdit, QPushButton, QComboBox,
    QDoubleSpinBox, QSpinBox, QTextEdit, QFileDialog, QMessageBox,
    QTabWidget, QCheckBox
)
from PySide6.QtCore import Qt, QTimer, Signal, QThread
from PySide6.QtGui import QFont, QPalette, QColor
import pyqtgraph as pg
import numpy as np


class DataAcquisitionThread(QThread):
    """Thread for acquiring data from hardware without blocking GUI"""
    data_ready = Signal(float, float)  # force, displacement
    
    def __init__(self):
        super().__init__()
        self.running = False
        self.test_type = "compression"
        self.target_displacement = 0.0
        self.current_displacement = 0.0
        self.current_force = 0.0
        
    def run(self):
        """Main acquisition loop - replace with actual hardware interface"""
        while self.running:
            # TODO: Replace with actual hardware reading
            # Example: read from load cell and displacement sensor
            # self.current_force = read_load_cell()
            # self.current_displacement = read_displacement_sensor()
            
            # Simulated data for testing
            self.current_displacement += 0.01
            self.current_force = 100 * np.sin(self.current_displacement) + np.random.normal(0, 5)
            
            self.data_ready.emit(self.current_force, self.current_displacement)
            self.msleep(50)  # 20 Hz update rate
            
    def stop(self):
        self.running = False


class TestBenchGUI(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Test Bench Control - Compression & Torsion")
        self.setGeometry(100, 100, 1200, 800)
        
        # Data storage
        self.force_data = []
        self.displacement_data = []
        self.time_data = []
        self.test_start_time = None
        self.test_running = False
        
        # Test parameters
        self.test_type = "compression"
        self.test_mode = "monotonic"
        self.target_displacement = 10.0
        self.max_cycles = 1000
        self.current_cycle = 0
        
        # Data acquisition thread
        self.daq_thread = DataAcquisitionThread()
        self.daq_thread.data_ready.connect(self.update_data)
        
        # Setup UI
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface"""
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        
        # Title
        title = QLabel("Test Bench Control System")
        title_font = QFont()
        title_font.setPointSize(18)
        title_font.setBold(True)
        title.setFont(title_font)
        title.setAlignment(Qt.AlignCenter)
        main_layout.addWidget(title)
        
        # Main content area with tabs
        tabs = QTabWidget()
        tabs.addTab(self.create_control_tab(), "Test Control")
        tabs.addTab(self.create_settings_tab(), "Settings")
        main_layout.addWidget(tabs)
        
        # Status bar
        self.status_label = QLabel("Status: Ready")
        self.status_label.setStyleSheet("padding: 5px; background-color: #e0e0e0;")
        main_layout.addWidget(self.status_label)
        
    def create_control_tab(self):
        """Create the main control tab"""
        tab = QWidget()
        layout = QHBoxLayout(tab)
        
        # Left panel - Controls
        left_panel = QVBoxLayout()
        
        # Test Configuration Group
        config_group = self.create_config_group()
        left_panel.addWidget(config_group)
        
        # Control Buttons Group
        control_group = self.create_control_group()
        left_panel.addWidget(control_group)
        
        # Current Readings Group
        readings_group = self.create_readings_group()
        left_panel.addWidget(readings_group)
        
        # Data Management Group
        data_group = self.create_data_group()
        left_panel.addWidget(data_group)
        
        left_panel.addStretch()
        layout.addLayout(left_panel, 1)
        
        # Right panel - Plot
        right_panel = QVBoxLayout()
        self.plot_widget = self.create_plot()
        right_panel.addWidget(self.plot_widget)
        layout.addLayout(right_panel, 2)
        
        return tab
    
    def create_config_group(self):
        """Create test configuration group"""
        group = QGroupBox("Test Configuration")
        layout = QVBoxLayout()
        
        # Test Type
        type_layout = QHBoxLayout()
        type_layout.addWidget(QLabel("Test Type:"))
        self.test_type_combo = QComboBox()
        self.test_type_combo.addItems(["Compression", "Torsion"])
        self.test_type_combo.currentTextChanged.connect(self.on_test_type_changed)
        type_layout.addWidget(self.test_type_combo)
        layout.addLayout(type_layout)
        
        # Test Mode
        mode_layout = QHBoxLayout()
        mode_layout.addWidget(QLabel("Test Mode:"))
        self.test_mode_combo = QComboBox()
        self.test_mode_combo.addItems(["Monotonic", "Fatigue"])
        self.test_mode_combo.currentTextChanged.connect(self.on_test_mode_changed)
        mode_layout.addWidget(self.test_mode_combo)
        layout.addLayout(mode_layout)
        
        # Target Displacement
        disp_layout = QHBoxLayout()
        disp_layout.addWidget(QLabel("Target Displacement (mm):"))
        self.displacement_spin = QDoubleSpinBox()
        self.displacement_spin.setRange(0.1, 100.0)
        self.displacement_spin.setValue(10.0)
        self.displacement_spin.setDecimals(2)
        disp_layout.addWidget(self.displacement_spin)
        layout.addLayout(disp_layout)
        
        # Max Cycles (for fatigue)
        cycles_layout = QHBoxLayout()
        cycles_layout.addWidget(QLabel("Max Cycles:"))
        self.cycles_spin = QSpinBox()
        self.cycles_spin.setRange(1, 1000000)
        self.cycles_spin.setValue(1000)
        self.cycles_spin.setEnabled(False)  # Disabled for monotonic
        cycles_layout.addWidget(self.cycles_spin)
        layout.addLayout(cycles_layout)
        
        group.setLayout(layout)
        return group
    
    def create_control_group(self):
        """Create control buttons group"""
        group = QGroupBox("Test Control")
        layout = QVBoxLayout()
        
        # Start Test Button
        self.start_button = QPushButton("START TEST")
        self.start_button.setStyleSheet("""
            QPushButton {
                background-color: #4CAF50;
                color: white;
                font-size: 16px;
                font-weight: bold;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.start_button.clicked.connect(self.start_test)
        layout.addWidget(self.start_button)
        
        # Stop Test Button
        self.stop_button = QPushButton("STOP TEST")
        self.stop_button.setStyleSheet("""
            QPushButton {
                background-color: #f44336;
                color: white;
                font-size: 16px;
                font-weight: bold;
                padding: 15px;
            }
            QPushButton:hover {
                background-color: #da190b;
            }
            QPushButton:disabled {
                background-color: #cccccc;
            }
        """)
        self.stop_button.clicked.connect(self.stop_test)
        self.stop_button.setEnabled(False)
        layout.addWidget(self.stop_button)
        
        # Emergency Stop Button
        self.emergency_button = QPushButton("EMERGENCY STOP")
        self.emergency_button.setStyleSheet("""
            QPushButton {
                background-color: #FF9800;
                color: white;
                font-size: 14px;
                font-weight: bold;
                padding: 10px;
            }
            QPushButton:hover {
                background-color: #F57C00;
            }
        """)
        self.emergency_button.clicked.connect(self.emergency_stop)
        layout.addWidget(self.emergency_button)
        
        group.setLayout(layout)
        return group
    
    def create_readings_group(self):
        """Create current readings display group"""
        group = QGroupBox("Current Readings")
        layout = QVBoxLayout()
        
        # Force reading
        force_layout = QHBoxLayout()
        force_layout.addWidget(QLabel("Force (N):"))
        self.force_display = QLabel("0.00")
        self.force_display.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                padding: 5px;
                background-color: #f0f0f0;
            }
        """)
        force_layout.addWidget(self.force_display)
        layout.addLayout(force_layout)
        
        # Displacement reading
        disp_layout = QHBoxLayout()
        disp_layout.addWidget(QLabel("Displacement (mm):"))
        self.disp_display = QLabel("0.00")
        self.disp_display.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                padding: 5px;
                background-color: #f0f0f0;
            }
        """)
        disp_layout.addWidget(self.disp_display)
        layout.addLayout(disp_layout)
        
        # Cycle count (for fatigue)
        cycle_layout = QHBoxLayout()
        cycle_layout.addWidget(QLabel("Cycle:"))
        self.cycle_display = QLabel("0")
        self.cycle_display.setStyleSheet("""
            QLabel {
                font-size: 18px;
                font-weight: bold;
                padding: 5px;
                background-color: #f0f0f0;
            }
        """)
        cycle_layout.addWidget(self.cycle_display)
        layout.addLayout(cycle_layout)
        
        group.setLayout(layout)
        return group
    
    def create_data_group(self):
        """Create data management group"""
        group = QGroupBox("Data Management")
        layout = QVBoxLayout()
        
        # Save Data Button
        save_button = QPushButton("Save Data (CSV)")
        save_button.clicked.connect(self.save_data_csv)
        layout.addWidget(save_button)
        
        # Save Excel Button
        save_excel_button = QPushButton("Save Data (Excel)")
        save_excel_button.clicked.connect(self.save_data_excel)
        layout.addWidget(save_excel_button)
        
        # Email Results Button
        email_button = QPushButton("Email Results")
        email_button.clicked.connect(self.email_results)
        layout.addWidget(email_button)
        
        # Clear Data Button
        clear_button = QPushButton("Clear Data")
        clear_button.clicked.connect(self.clear_data)
        layout.addWidget(clear_button)
        
        group.setLayout(layout)
        return group
    
    def create_plot(self):
        """Create the real-time plot widget"""
        plot_widget = pg.PlotWidget()
        plot_widget.setBackground('w')
        plot_widget.setLabel('left', 'Force (N)')
        plot_widget.setLabel('bottom', 'Displacement (mm)')
        plot_widget.setTitle('Force vs Displacement', color='k', size='14pt')
        plot_widget.showGrid(x=True, y=True, alpha=0.3)
        
        # Create plot curve
        self.curve = plot_widget.plot(pen=pg.mkPen(color='b', width=2))
        
        return plot_widget
    
    def create_settings_tab(self):
        """Create settings tab"""
        tab = QWidget()
        layout = QVBoxLayout(tab)
        
        # Email Settings Group
        email_group = QGroupBox("Email Settings")
        email_layout = QVBoxLayout()
        
        # Email address
        email_addr_layout = QHBoxLayout()
        email_addr_layout.addWidget(QLabel("Your Email:"))
        self.email_input = QLineEdit()
        self.email_input.setPlaceholderText("your.email@example.com")
        email_addr_layout.addWidget(self.email_input)
        email_layout.addLayout(email_addr_layout)
        
        # SMTP Server
        smtp_layout = QHBoxLayout()
        smtp_layout.addWidget(QLabel("SMTP Server:"))
        self.smtp_input = QLineEdit()
        self.smtp_input.setPlaceholderText("smtp.gmail.com")
        self.smtp_input.setText("smtp.gmail.com")
        smtp_layout.addWidget(self.smtp_input)
        email_layout.addLayout(smtp_layout)
        
        # SMTP Port
        port_layout = QHBoxLayout()
        port_layout.addWidget(QLabel("SMTP Port:"))
        self.smtp_port_input = QSpinBox()
        self.smtp_port_input.setRange(1, 65535)
        self.smtp_port_input.setValue(587)
        port_layout.addWidget(self.smtp_port_input)
        email_layout.addLayout(port_layout)
        
        # SMTP Password
        password_layout = QHBoxLayout()
        password_layout.addWidget(QLabel("SMTP Password:"))
        self.smtp_password_input = QLineEdit()
        self.smtp_password_input.setEchoMode(QLineEdit.Password)
        self.smtp_password_input.setPlaceholderText("App-specific password")
        password_layout.addWidget(self.smtp_password_input)
        email_layout.addLayout(password_layout)
        
        email_group.setLayout(email_layout)
        layout.addWidget(email_group)
        
        # Hardware Settings Group
        hardware_group = QGroupBox("Hardware Settings")
        hardware_layout = QVBoxLayout()
        hardware_layout.addWidget(QLabel("Configure hardware interface here"))
        # TODO: Add hardware configuration options
        hardware_group.setLayout(hardware_layout)
        layout.addWidget(hardware_group)
        
        layout.addStretch()
        return tab
    
    def on_test_type_changed(self, text):
        """Handle test type change"""
        self.test_type = text.lower()
        self.update_status(f"Test type changed to: {text}")
    
    def on_test_mode_changed(self, text):
        """Handle test mode change"""
        self.test_mode = text.lower()
        # Enable/disable max cycles for fatigue mode
        if text.lower() == "fatigue":
            self.cycles_spin.setEnabled(True)
        else:
            self.cycles_spin.setEnabled(False)
        self.update_status(f"Test mode changed to: {text}")
    
    def start_test(self):
        """Start the test"""
        if self.test_running:
            return
        
        # Get parameters
        self.target_displacement = self.displacement_spin.value()
        self.max_cycles = self.cycles_spin.value()
        
        # Reset data
        self.force_data = []
        self.displacement_data = []
        self.time_data = []
        self.current_cycle = 0
        self.test_start_time = datetime.datetime.now()
        
        # Update UI
        self.test_running = True
        self.start_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.test_type_combo.setEnabled(False)
        self.test_mode_combo.setEnabled(False)
        self.displacement_spin.setEnabled(False)
        self.cycles_spin.setEnabled(False)
        
        # Start data acquisition
        self.daq_thread.test_type = self.test_type
        self.daq_thread.target_displacement = self.target_displacement
        self.daq_thread.running = True
        self.daq_thread.start()
        
        self.update_status(f"Test started: {self.test_type.capitalize()} - {self.test_mode.capitalize()}")
    
    def stop_test(self):
        """Stop the test normally"""
        if not self.test_running:
            return
        
        # Stop data acquisition
        self.daq_thread.stop()
        self.daq_thread.wait()
        
        # Update UI
        self.test_running = False
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.test_type_combo.setEnabled(True)
        self.test_mode_combo.setEnabled(True)
        self.displacement_spin.setEnabled(True)
        if self.test_mode == "fatigue":
            self.cycles_spin.setEnabled(True)
        
        self.update_status("Test stopped")
    
    def emergency_stop(self):
        """Emergency stop - immediately halt everything"""
        if self.test_running:
            self.daq_thread.stop()
            self.daq_thread.wait()
        
        # TODO: Send emergency stop signal to hardware
        
        # Reset UI
        self.test_running = False
        self.start_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.test_type_combo.setEnabled(True)
        self.test_mode_combo.setEnabled(True)
        self.displacement_spin.setEnabled(True)
        if self.test_mode == "fatigue":
            self.cycles_spin.setEnabled(True)
        
        self.update_status("EMERGENCY STOP ACTIVATED", error=True)
        QMessageBox.warning(self, "Emergency Stop", "Emergency stop activated!")
    
    def update_data(self, force, displacement):
        """Update data from acquisition thread"""
        if not self.test_running:
            return
        
        # Store data
        elapsed_time = (datetime.datetime.now() - self.test_start_time).total_seconds()
        self.force_data.append(force)
        self.displacement_data.append(displacement)
        self.time_data.append(elapsed_time)
        
        # Update displays
        self.force_display.setText(f"{force:.2f}")
        self.disp_display.setText(f"{displacement:.3f}")
        
        # Update plot
        self.curve.setData(self.displacement_data, self.force_data)
        
        # Check test completion
        if self.test_mode == "monotonic":
            if displacement >= self.target_displacement:
                self.stop_test()
                self.update_status("Test completed!")
        else:  # fatigue
            # TODO: Implement cycle detection logic
            pass
    
    def save_data_csv(self):
        """Save data to CSV file"""
        if len(self.force_data) == 0:
            QMessageBox.warning(self, "No Data", "No data to save!")
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save CSV File", "", "CSV Files (*.csv)"
        )
        
        if filename:
            try:
                with open(filename, 'w', newline='') as csvfile:
                    writer = csv.writer(csvfile)
                    # Header
                    writer.writerow([
                        f"Test Type: {self.test_type}",
                        f"Test Mode: {self.test_mode}",
                        f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                    ])
                    writer.writerow([])
                    writer.writerow(['Time (s)', 'Displacement (mm)', 'Force (N)'])
                    # Data
                    for t, d, f in zip(self.time_data, self.displacement_data, self.force_data):
                        writer.writerow([f"{t:.3f}", f"{d:.3f}", f"{f:.2f}"])
                
                self.update_status(f"Data saved to {filename}")
                QMessageBox.information(self, "Success", f"Data saved to:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")
    
    def save_data_excel(self):
        """Save data to Excel file"""
        if len(self.force_data) == 0:
            QMessageBox.warning(self, "No Data", "No data to save!")
            return
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl package is required for Excel export.\n"
                "Install it with: pip install openpyxl"
            )
            return
        
        filename, _ = QFileDialog.getSaveFileName(
            self, "Save Excel File", "", "Excel Files (*.xlsx)"
        )
        
        if filename:
            try:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Test Data"
                
                # Header information
                ws['A1'] = "Test Bench Results"
                ws['A1'].font = Font(bold=True, size=14)
                ws['A2'] = f"Test Type: {self.test_type.capitalize()}"
                ws['A3'] = f"Test Mode: {self.test_mode.capitalize()}"
                ws['A4'] = f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                
                # Column headers
                header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                headers = ['Time (s)', 'Displacement (mm)', 'Force (N)']
                for col, header in enumerate(headers, start=1):
                    cell = ws.cell(row=6, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                
                # Data
                for row, (t, d, f) in enumerate(zip(self.time_data, self.displacement_data, self.force_data), start=7):
                    ws.cell(row=row, column=1, value=float(f"{t:.3f}"))
                    ws.cell(row=row, column=2, value=float(f"{d:.3f}"))
                    ws.cell(row=row, column=3, value=float(f"{f:.2f}"))
                
                wb.save(filename)
                self.update_status(f"Data saved to {filename}")
                QMessageBox.information(self, "Success", f"Data saved to:\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save file:\n{str(e)}")
    
    def email_results(self):
        """Email test results"""
        if len(self.force_data) == 0:
            QMessageBox.warning(self, "No Data", "No data to email!")
            return
        
        email = self.email_input.text()
        smtp_server = self.smtp_input.text()
        smtp_port = self.smtp_port_input.value()
        password = self.smtp_password_input.text()
        
        if not email or not smtp_server or not password:
            QMessageBox.warning(
                self, "Missing Information",
                "Please fill in email settings in the Settings tab."
            )
            return
        
        try:
            # Create temporary CSV file
            temp_filename = f"test_results_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            temp_path = Path.home() / temp_filename
            
            with open(temp_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow([
                    f"Test Type: {self.test_type}",
                    f"Test Mode: {self.test_mode}",
                    f"Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                ])
                writer.writerow([])
                writer.writerow(['Time (s)', 'Displacement (mm)', 'Force (N)'])
                for t, d, f in zip(self.time_data, self.displacement_data, self.force_data):
                    writer.writerow([f"{t:.3f}", f"{d:.3f}", f"{f:.2f}"])
            
            # Create email
            msg = MIMEMultipart()
            msg['From'] = email
            msg['To'] = email
            msg['Subject'] = f"Test Bench Results - {self.test_type.capitalize()} {self.test_mode.capitalize()}"
            
            body = f"""
Test Bench Results

Test Type: {self.test_type.capitalize()}
Test Mode: {self.test_mode.capitalize()}
Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

Data points collected: {len(self.force_data)}
Max Force: {max(self.force_data):.2f} N
Max Displacement: {max(self.displacement_data):.3f} mm

See attached CSV file for full data.
            """
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach CSV file
            with open(temp_path, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header('Content-Disposition', f'attachment; filename= {temp_filename}')
            msg.attach(part)
            
            # Send email
            with smtplib.SMTP(smtp_server, smtp_port) as server:
                server.starttls()
                server.login(email, password)
                server.send_message(msg)
            
            # Clean up temp file
            temp_path.unlink()
            
            self.update_status("Results emailed successfully!")
            QMessageBox.information(self, "Success", "Results emailed successfully!")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to send email:\n{str(e)}")
    
    def clear_data(self):
        """Clear all collected data"""
        reply = QMessageBox.question(
            self, "Clear Data",
            "Are you sure you want to clear all data?",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            self.force_data = []
            self.displacement_data = []
            self.time_data = []
            self.current_cycle = 0
            self.curve.setData([], [])
            self.force_display.setText("0.00")
            self.disp_display.setText("0.00")
            self.cycle_display.setText("0")
            self.update_status("Data cleared")
    
    def update_status(self, message, error=False):
        """Update status bar message"""
        if error:
            self.status_label.setStyleSheet("padding: 5px; background-color: #ffcccc; color: #cc0000;")
        else:
            self.status_label.setStyleSheet("padding: 5px; background-color: #e0e0e0;")
        self.status_label.setText(f"Status: {message}")


def main():
    app = QApplication(sys.argv)
    window = TestBenchGUI()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
